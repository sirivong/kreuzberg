"""Security edge-case fixtures.

Five fixtures exercising the OOXML-extractor guards:

- ``xlsx_dde_formula.xlsx`` — workbook carrying ``=HYPERLINK(…)`` and
  ``=DDE(…)`` formula calls. GT asserts extraction succeeds and surfaces
  a warning naming the dangerous formula type.
- ``xlsx_safe_formulas.xlsx`` — control workbook with only ``=SUM(A1:A2)``.
  GT asserts no DDE / HYPERLINK warning.
- ``docx_oversized_embedded.docx`` — a DOCX whose embedded part is
  declared at 100 MiB (a synthetic stream of zeros). GT asserts that with
  ``max_embedded_file_bytes = 50 MiB`` the extractor skips the child and
  emits a size-limit warning.
- ``zip_bomb_xlsx.xlsx`` — 50:1 compression ratio. GT asserts extraction
  SUCCEEDS — the guard tolerates legitimate compression.
- ``zip_bomb_xlsx_pathological.xlsx`` — 200:1 ratio. GT asserts extraction
  is REJECTED by the zip-bomb guard.

All zip archives use a fixed mtime for hash-stable output.
"""

from __future__ import annotations

import io
import zipfile
from pathlib import Path

from openpyxl import Workbook  # type: ignore[import-untyped, import-not-found, unused-ignore]

from .gt_schema import security_expectation, write_ground_truth

ZIP_MTIME = (2024, 1, 1, 0, 0, 0)

ONE_MIB = 1024 * 1024
OVERSIZED_BYTES = 100 * ONE_MIB

SAFE_COMPRESSION_RATIO = 50
PATHOLOGICAL_COMPRESSION_RATIO = 200

COMPRESSED_ENTRY_BYTES = 64 * 1024


def _rewrite_zip(src_bytes: bytes, additions: dict[str, bytes], replacements: dict[str, bytes]) -> bytes:
    """Re-zip ``src_bytes`` with deterministic mtimes; additions appended."""
    buf = io.BytesIO()
    seen: set[str] = set()
    with zipfile.ZipFile(io.BytesIO(src_bytes), "r") as src:
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as dst:
            for name in src.namelist():
                data = replacements.get(name, src.read(name))
                info = zipfile.ZipInfo(name, ZIP_MTIME)
                info.compress_type = zipfile.ZIP_DEFLATED
                dst.writestr(info, data)
                seen.add(name)
            for name, data in additions.items():
                if name in seen:
                    continue
                info = zipfile.ZipInfo(name, ZIP_MTIME)
                info.compress_type = zipfile.ZIP_DEFLATED
                dst.writestr(info, data)
    return buf.getvalue()


def _emit_xlsx_dde(output_dir: Path, repo_root: Path) -> list[Path]:
    """Workbook with one HYPERLINK and one DDE formula cell."""
    wb = Workbook()
    ws = wb.active
    ws.title = "danger"
    ws["A1"] = "label"
    ws["A2"] = "ok"
    ws["B1"] = '=HYPERLINK("https://example.com/evil", "click me")'
    ws["B2"] = '=DDE("cmd","/c calc.exe","_")'
    buf = io.BytesIO()
    wb.save(buf)

    fixture_path = output_dir / "xlsx_dde_formula.xlsx"
    sidecar_path = output_dir / "xlsx_dde_formula.gt.json"
    fixture_path.write_bytes(buf.getvalue())
    write_ground_truth(
        sidecar_path,
        fixture_path,
        repo_root,
        document_format="xlsx",
        feature="security",
        expectations=security_expectation(
            should_extract=True,
            expected_warnings=["dde", "hyperlink"],
            notes=(
                "Two dangerous formula calls in B1/B2. Extraction must succeed but the "
                "warnings stream must mention DDE and HYPERLINK (case-insensitive substring "
                "match — exact wording is up to the extractor's audit emitter)."
            ),
        ),
        generator="security_fixtures",
    )
    return [fixture_path, sidecar_path]


def _emit_xlsx_safe(output_dir: Path, repo_root: Path) -> list[Path]:
    """Control workbook with only a SUM formula — no warnings expected."""
    wb = Workbook()
    ws = wb.active
    ws.title = "safe"
    ws["A1"] = 10
    ws["A2"] = 32
    ws["A3"] = "=SUM(A1:A2)"
    buf = io.BytesIO()
    wb.save(buf)

    fixture_path = output_dir / "xlsx_safe_formulas.xlsx"
    sidecar_path = output_dir / "xlsx_safe_formulas.gt.json"
    fixture_path.write_bytes(buf.getvalue())
    write_ground_truth(
        sidecar_path,
        fixture_path,
        repo_root,
        document_format="xlsx",
        feature="security",
        expectations=security_expectation(
            should_extract=True,
            expected_warnings=[],
            notes=(
                "Control workbook. Asserts the DDE/HYPERLINK warning path does NOT trigger on "
                "ordinary arithmetic formulas — guards against false positives."
            ),
        ),
        generator="security_fixtures",
    )
    return [fixture_path, sidecar_path]


def _emit_docx_oversized_embedded(output_dir: Path, repo_root: Path) -> list[Path]:
    """DOCX whose ``word/embeddings/oversized.bin`` is a 100 MiB zero stream."""
    from docx import Document  # type: ignore[import-untyped, import-not-found, unused-ignore]

    doc = Document()
    doc.add_paragraph("Document carrying an oversized embedded part.")
    base_buf = io.BytesIO()
    doc.save(base_buf)
    base_bytes = base_buf.getvalue()

    oversized_payload = b"\x00" * OVERSIZED_BYTES

    with zipfile.ZipFile(io.BytesIO(base_bytes), "r") as zf:
        content_types = zf.read("[Content_Types].xml")
        document_rels = zf.read("word/_rels/document.xml.rels")

    new_content_types = content_types.replace(
        b"</Types>",
        b'<Override PartName="/word/embeddings/oversized.bin" ContentType="application/octet-stream"/></Types>',
    )
    new_document_rels = document_rels.replace(
        b"</Relationships>",
        b'<Relationship Id="rIdOversized" '
        b'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/oleObject" '
        b'Target="embeddings/oversized.bin"/></Relationships>',
    )

    out_bytes = _rewrite_zip(
        base_bytes,
        additions={"word/embeddings/oversized.bin": oversized_payload},
        replacements={
            "[Content_Types].xml": new_content_types,
            "word/_rels/document.xml.rels": new_document_rels,
        },
    )

    fixture_path = output_dir / "docx_oversized_embedded.docx"
    sidecar_path = output_dir / "docx_oversized_embedded.gt.json"
    fixture_path.write_bytes(out_bytes)
    write_ground_truth(
        sidecar_path,
        fixture_path,
        repo_root,
        document_format="docx",
        feature="security",
        expectations=security_expectation(
            should_extract=True,
            expected_warnings=["embed", "size", "skip"],
            notes=(
                "word/embeddings/oversized.bin carries 100 MiB of zeros. With "
                "max_embedded_file_bytes = 50 MiB the extractor must skip the child "
                "and emit a warning mentioning the embed + size + skip. The base "
                "document is extracted normally."
            ),
        ),
        generator="security_fixtures",
    )
    return [fixture_path, sidecar_path]


def _build_zip_bomb_xlsx(compression_ratio: int) -> bytes:
    """Author an XLSX whose embedded /xl/payload.bin has the requested ratio.

    Implemented by writing ``compression_ratio * COMPRESSED_ENTRY_BYTES``
    bytes of zeros into a part that compresses down to roughly
    ``COMPRESSED_ENTRY_BYTES``. The XLSX shell is otherwise a valid one-
    sheet workbook so the zip-bomb guard is what triggers (or doesn't),
    not a malformed-archive code path.
    """
    wb = Workbook()
    ws = wb.active
    ws["A1"] = f"Compression ratio target: {compression_ratio}:1"
    base_buf = io.BytesIO()
    wb.save(base_buf)
    base_bytes = base_buf.getvalue()

    uncompressed_size = compression_ratio * COMPRESSED_ENTRY_BYTES
    payload = b"\x00" * uncompressed_size

    return _rewrite_zip(
        base_bytes,
        additions={"xl/payload.bin": payload},
        replacements={},
    )


def _emit_zip_bomb_pair(output_dir: Path, repo_root: Path) -> list[Path]:
    written: list[Path] = []

    safe_path = output_dir / "zip_bomb_xlsx.xlsx"
    safe_sidecar = output_dir / "zip_bomb_xlsx.gt.json"
    safe_path.write_bytes(_build_zip_bomb_xlsx(SAFE_COMPRESSION_RATIO))
    write_ground_truth(
        safe_sidecar,
        safe_path,
        repo_root,
        document_format="xlsx",
        feature="security",
        expectations=security_expectation(
            should_extract=True,
            expected_warnings=[],
            notes=(
                f"{SAFE_COMPRESSION_RATIO}:1 compression ratio — legitimately compressible "
                "content (zero-filled stream). The zip-bomb guard must NOT trigger; this "
                "fixture verifies the guard tolerates real-world compression."
            ),
        ),
        generator="security_fixtures",
    )
    written.extend([safe_path, safe_sidecar])

    pathological_path = output_dir / "zip_bomb_xlsx_pathological.xlsx"
    pathological_sidecar = output_dir / "zip_bomb_xlsx_pathological.gt.json"
    pathological_path.write_bytes(_build_zip_bomb_xlsx(PATHOLOGICAL_COMPRESSION_RATIO))
    write_ground_truth(
        pathological_sidecar,
        pathological_path,
        repo_root,
        document_format="xlsx",
        feature="security",
        expectations=security_expectation(
            should_extract=False,
            expected_warnings=["zip", "bomb"],
            notes=(
                f"{PATHOLOGICAL_COMPRESSION_RATIO}:1 compression ratio — the zip-bomb guard "
                "MUST reject the file. Extraction returns an error whose message mentions "
                "zip/bomb (case-insensitive substring match)."
            ),
        ),
        generator="security_fixtures",
    )
    written.extend([pathological_path, pathological_sidecar])
    return written


def generate(output_root: Path, repo_root: Path) -> list[Path]:
    """Emit all security fixtures under ``output_root/security/``."""
    output_dir = output_root / "security"
    output_dir.mkdir(parents=True, exist_ok=True)
    written: list[Path] = []
    written.extend(_emit_xlsx_dde(output_dir, repo_root))
    written.extend(_emit_xlsx_safe(output_dir, repo_root))
    written.extend(_emit_docx_oversized_embedded(output_dir, repo_root))
    written.extend(_emit_zip_bomb_pair(output_dir, repo_root))
    return written
